"""Sanctions list ingestion from official government sources.

Downloads and parses sanctions data from:
- OFAC SDN List (Treasury) - XML format
- OFAC Consolidated Sanctions List - XML format
- BIS Entity List (Commerce) - Excel format
- BIS Denied Persons List (Commerce) - TXT format
"""

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import Element  # Type only

import defusedxml.ElementTree as ET  # noqa: N817 - ET is standard convention
import httpx

from export_control_mcp.models.sanctions import (
    DeniedPersonEntry,
    EntityListEntry,
    EntityType,
    SDNEntry,
)
from export_control_mcp.services.sanctions_db import SanctionsDBService

logger = logging.getLogger(__name__)


# Official source URLs
OFAC_SOURCES = {
    "sdn_xml": "https://www.treasury.gov/ofac/downloads/sdn.xml",
    "sdn_csv": "https://www.treasury.gov/ofac/downloads/sdn.csv",
    "consolidated_xml": "https://www.treasury.gov/ofac/downloads/consolidated/consolidated.xml",
    "consolidated_csv": "https://www.treasury.gov/ofac/downloads/consolidated/consolidated.csv",
}

BIS_SOURCES = {
    "denied_persons_txt": "https://www.bis.doc.gov/dpl/dpl.txt",
    "entity_list_page": "https://www.bis.doc.gov/index.php/policy-guidance/lists-of-parties-of-concern/entity-list",
    # Note: Entity List Excel file URL changes; may need to scrape the page
}


class SanctionsIngestor:
    """Ingest sanctions lists from official government sources."""

    def __init__(self, db: SanctionsDBService, download_dir: Path | None = None):
        """
        Initialize the sanctions ingestor.

        Args:
            db: Sanctions database service for storing entries.
            download_dir: Directory for downloaded files (default: ./data/downloads)
        """
        self.db = db
        self.download_dir = download_dir or Path("./data/downloads")
        self.download_dir.mkdir(parents=True, exist_ok=True)

    async def download_file(self, url: str, filename: str) -> Path | None:
        """
        Download a file from a URL.

        Args:
            url: URL to download from.
            filename: Name for the downloaded file.

        Returns:
            Path to downloaded file, or None if download failed.
        """
        output_path = self.download_dir / filename

        try:
            async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
                logger.info(f"Downloading {url}...")
                response = await client.get(url)
                response.raise_for_status()

                output_path.write_bytes(response.content)
                logger.info(f"Downloaded to {output_path} ({len(response.content)} bytes)")
                return output_path

        except Exception as e:
            logger.error(f"Failed to download {url}: {e}")
            return None

    # =========================================================================
    # OFAC SDN List Ingestion
    # =========================================================================

    async def ingest_ofac_sdn(self, force_download: bool = False) -> dict[str, Any]:
        """
        Download and ingest the OFAC SDN List.

        Args:
            force_download: If True, download even if file exists.

        Returns:
            Dictionary with ingestion statistics.
        """
        result = {
            "source": "OFAC SDN List",
            "url": OFAC_SOURCES["sdn_xml"],
            "entries_added": 0,
            "entries_skipped": 0,
            "errors": [],
        }

        # Download the XML file
        xml_path = self.download_dir / "sdn.xml"
        if force_download or not xml_path.exists():
            xml_path = await self.download_file(OFAC_SOURCES["sdn_xml"], "sdn.xml")
            if not xml_path:
                result["errors"].append("Failed to download SDN XML")
                return result

        # Parse the XML
        try:
            entries = self._parse_ofac_sdn_xml(xml_path)
            logger.info(f"Parsed {len(entries)} SDN entries from XML")

            for entry in entries:
                try:
                    self.db.add_sdn_entry(entry)
                    result["entries_added"] += 1
                except Exception as e:
                    result["entries_skipped"] += 1
                    if len(result["errors"]) < 10:  # Limit error messages
                        result["errors"].append(f"Error adding {entry.name}: {e}")

        except Exception as e:
            result["errors"].append(f"XML parsing error: {e}")
            logger.error(f"Failed to parse SDN XML: {e}")

        return result

    def _parse_ofac_sdn_xml(self, xml_path: Path) -> list[SDNEntry]:
        """Parse OFAC SDN XML file into SDNEntry objects."""
        entries = []
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Find all SDN entries (try with and without namespace)
        # OFAC changed namespace in 2024+ to sanctionslistservice.ofac.treas.gov
        sdn_entries = root.findall(
            ".//{https://sanctionslistservice.ofac.treas.gov/api/PublicationPreview/exports/XML}sdnEntry"
        )
        if not sdn_entries:
            # Try without namespace (older format)
            sdn_entries = root.findall(".//sdnEntry")

        # Define namespace prefix for element lookups
        ns_prefix = (
            "{https://sanctionslistservice.ofac.treas.gov/api/PublicationPreview/exports/XML}"
        )

        for sdn in sdn_entries:
            try:
                # Extract basic info (try with namespace first, then without)
                uid = self._get_text(sdn, f"{ns_prefix}uid") or self._get_text(sdn, "uid")
                name = self._get_text(sdn, f"{ns_prefix}lastName") or self._get_text(
                    sdn, "lastName"
                )
                first_name = self._get_text(sdn, f"{ns_prefix}firstName") or self._get_text(
                    sdn, "firstName"
                )
                sdn_type = self._get_text(sdn, f"{ns_prefix}sdnType") or self._get_text(
                    sdn, "sdnType"
                )

                if not uid or not name:
                    continue

                # Combine first and last name for individuals
                full_name = f"{name}, {first_name}" if first_name else name

                # Map SDN type
                entity_type = self._map_sdn_type(sdn_type or "")

                # Extract programs
                programs = []
                program_list = sdn.find(f"{ns_prefix}programList") or sdn.find("programList")
                if program_list is not None:
                    for prog in program_list.findall(f"{ns_prefix}program") or program_list.findall(
                        "program"
                    ):
                        if prog.text:
                            programs.append(prog.text.strip())

                # Extract aliases
                aliases = []
                aka_list = sdn.find(f"{ns_prefix}akaList") or sdn.find("akaList")
                if aka_list is not None:
                    for aka in aka_list.findall(f"{ns_prefix}aka") or aka_list.findall("aka"):
                        aka_name = self._get_text(aka, f"{ns_prefix}lastName") or self._get_text(
                            aka, "lastName"
                        )
                        if aka_name:
                            aliases.append(aka_name)

                # Extract addresses
                addresses = []
                addr_list = sdn.find(f"{ns_prefix}addressList") or sdn.find("addressList")
                if addr_list is not None:
                    for addr in addr_list.findall(f"{ns_prefix}address") or addr_list.findall(
                        "address"
                    ):
                        addr_parts = []
                        for field in [
                            "address1",
                            "address2",
                            "address3",
                            "city",
                            "stateOrProvince",
                            "country",
                        ]:
                            val = self._get_text(addr, f"{ns_prefix}{field}") or self._get_text(
                                addr, field
                            )
                            if val:
                                addr_parts.append(val)
                        if addr_parts:
                            addresses.append(", ".join(addr_parts))

                # Extract nationalities and DOBs for individuals
                nationalities = []
                dates_of_birth = []
                id_list = sdn.find(f"{ns_prefix}idList") or sdn.find("idList")
                if id_list is not None:
                    for id_elem in id_list.findall(f"{ns_prefix}id") or id_list.findall("id"):
                        id_type = self._get_text(id_elem, f"{ns_prefix}idType") or self._get_text(
                            id_elem, "idType"
                        )
                        id_num = self._get_text(id_elem, f"{ns_prefix}idNumber") or self._get_text(
                            id_elem, "idNumber"
                        )
                        if id_type and id_num and "nationality" in id_type.lower():
                            nationalities.append(id_num)

                # Extract remarks
                remarks = (
                    self._get_text(sdn, f"{ns_prefix}remarks")
                    or self._get_text(sdn, "remarks")
                    or ""
                )

                entry = SDNEntry(
                    id=f"SDN-{uid}",
                    name=full_name,
                    sdn_type=entity_type,
                    programs=programs,
                    aliases=aliases,
                    addresses=addresses,
                    nationalities=nationalities,
                    dates_of_birth=dates_of_birth,
                    remarks=remarks,
                )
                entries.append(entry)

            except Exception as e:
                logger.warning(f"Error parsing SDN entry: {e}")
                continue

        return entries

    def _get_text(self, elem: Element, tag: str, ns: dict | None = None) -> str | None:
        """Get text content of a child element."""
        if ns:
            child = elem.find(tag, ns)
        else:
            child = elem.find(tag)
        return child.text.strip() if child is not None and child.text else None

    def _map_sdn_type(self, sdn_type: str) -> EntityType:
        """Map OFAC SDN type to our EntityType enum."""
        sdn_type_lower = sdn_type.lower()
        if "individual" in sdn_type_lower:
            return EntityType.INDIVIDUAL
        elif "vessel" in sdn_type_lower:
            return EntityType.VESSEL
        elif "aircraft" in sdn_type_lower:
            return EntityType.AIRCRAFT
        else:
            return EntityType.ENTITY

    # =========================================================================
    # BIS Denied Persons List Ingestion
    # =========================================================================

    async def ingest_bis_denied_persons(self, force_download: bool = False) -> dict[str, Any]:
        """
        Download and ingest the BIS Denied Persons List.

        Args:
            force_download: If True, download even if file exists.

        Returns:
            Dictionary with ingestion statistics.
        """
        result = {
            "source": "BIS Denied Persons List",
            "url": BIS_SOURCES["denied_persons_txt"],
            "entries_added": 0,
            "entries_skipped": 0,
            "errors": [],
        }

        # Download the TXT file
        txt_path = self.download_dir / "dpl.txt"
        if force_download or not txt_path.exists():
            txt_path = await self.download_file(BIS_SOURCES["denied_persons_txt"], "dpl.txt")
            if not txt_path:
                result["errors"].append("Failed to download Denied Persons List")
                return result

        # Parse the TXT file
        try:
            entries = self._parse_bis_denied_persons_txt(txt_path)
            logger.info(f"Parsed {len(entries)} Denied Persons entries")

            for entry in entries:
                try:
                    self.db.add_denied_person(entry)
                    result["entries_added"] += 1
                except Exception as e:
                    result["entries_skipped"] += 1
                    if len(result["errors"]) < 10:
                        result["errors"].append(f"Error adding {entry.name}: {e}")

        except Exception as e:
            result["errors"].append(f"TXT parsing error: {e}")
            logger.error(f"Failed to parse Denied Persons TXT: {e}")

        return result

    def _parse_bis_denied_persons_txt(self, txt_path: Path) -> list[DeniedPersonEntry]:
        """Parse BIS Denied Persons List TXT file."""
        entries = []
        content = txt_path.read_text(encoding="utf-8", errors="replace")

        # The DPL format is pipe-delimited or tab-delimited
        lines = content.strip().split("\n")

        for i, line in enumerate(lines):
            if not line.strip():
                continue

            # Skip header lines
            if line.startswith("#") or "NAME" in line.upper() and i < 5:
                continue

            # Try different delimiters
            parts = None
            for delimiter in ["|", "\t", ","]:
                parts = line.split(delimiter)
                if len(parts) >= 3:
                    break

            if not parts or len(parts) < 2:
                continue

            try:
                # Typical format: Name | Address | Effective Date | Expiration Date | FR Citation
                name = parts[0].strip()
                if not name:
                    continue

                addresses = [parts[1].strip()] if len(parts) > 1 and parts[1].strip() else []

                # Parse dates
                effective_date = None
                expiration_date = None
                fr_citation = ""

                if len(parts) > 2:
                    effective_date = self._parse_date(parts[2].strip())
                if len(parts) > 3:
                    expiration_date = self._parse_date(parts[3].strip())
                if len(parts) > 4:
                    fr_citation = parts[4].strip()

                entry = DeniedPersonEntry(
                    id=f"DP-{i:05d}",
                    name=name,
                    addresses=addresses,
                    effective_date=effective_date,
                    expiration_date=expiration_date,
                    federal_register_citation=fr_citation,
                )
                entries.append(entry)

            except Exception as e:
                logger.warning(f"Error parsing DPL line {i}: {e}")
                continue

        return entries

    def _parse_date(self, date_str: str) -> date | None:
        """Parse a date string in various formats."""
        if not date_str:
            return None

        formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%m-%d-%Y",
            "%d-%b-%Y",
            "%B %d, %Y",
            "%b %d, %Y",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    # =========================================================================
    # BIS Entity List Ingestion
    # =========================================================================

    async def ingest_bis_entity_list(
        self,
        excel_path: Path | None = None,
        force_download: bool = False,
    ) -> dict[str, Any]:
        """
        Ingest the BIS Entity List from an Excel file.

        Note: The Entity List Excel URL changes frequently. You may need to
        download it manually from the BIS website or provide the path.

        Args:
            excel_path: Path to downloaded Entity List Excel file.
            force_download: If True, attempt to download (may fail if URL changed).

        Returns:
            Dictionary with ingestion statistics.
        """
        result = {
            "source": "BIS Entity List",
            "url": BIS_SOURCES["entity_list_page"],
            "entries_added": 0,
            "entries_skipped": 0,
            "errors": [],
        }

        if not excel_path:
            result["errors"].append(
                "Entity List Excel path required. Download from: "
                "https://www.bis.doc.gov/index.php/policy-guidance/lists-of-parties-of-concern/entity-list"
            )
            return result

        if not excel_path.exists():
            result["errors"].append(f"File not found: {excel_path}")
            return result

        # Parse the Excel file
        try:
            entries = self._parse_bis_entity_list_excel(excel_path)
            logger.info(f"Parsed {len(entries)} Entity List entries")

            for entry in entries:
                try:
                    self.db.add_entity_list_entry(entry)
                    result["entries_added"] += 1
                except Exception as e:
                    result["entries_skipped"] += 1
                    if len(result["errors"]) < 10:
                        result["errors"].append(f"Error adding {entry.name}: {e}")

        except Exception as e:
            result["errors"].append(f"Excel parsing error: {e}")
            logger.error(f"Failed to parse Entity List Excel: {e}")

        return result

    def _parse_bis_entity_list_excel(self, excel_path: Path) -> list[EntityListEntry]:
        """Parse BIS Entity List Excel file."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return []

        entries = []
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active

        if ws is None:
            return entries

        # Find header row and column indices
        headers = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cell_lower = cell.lower()
                    if "name" in cell_lower and "alias" not in cell_lower:
                        headers["name"] = col_idx
                    elif "alias" in cell_lower:
                        headers["aliases"] = col_idx
                    elif "address" in cell_lower:
                        headers["address"] = col_idx
                    elif "country" in cell_lower:
                        headers["country"] = col_idx
                    elif "license requirement" in cell_lower:
                        headers["license_req"] = col_idx
                    elif "license policy" in cell_lower:
                        headers["license_policy"] = col_idx
                    elif "federal register" in cell_lower:
                        headers["fr_citation"] = col_idx
                    elif "effective date" in cell_lower or "date" in cell_lower:
                        headers["effective_date"] = col_idx

            if "name" in headers:
                # Found headers, start from next row
                data_start_row = row_idx + 1
                break
        else:
            logger.error("Could not find header row in Entity List Excel")
            return entries

        # Parse data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True)):
            try:
                row_data = list(row)

                name = (
                    row_data[headers["name"] - 1]
                    if "name" in headers and len(row_data) >= headers["name"]
                    else None
                )
                if not name:
                    continue

                aliases_str = (
                    row_data[headers["aliases"] - 1]
                    if "aliases" in headers and len(row_data) >= headers["aliases"]
                    else ""
                )
                aliases = [a.strip() for a in str(aliases_str or "").split(";") if a.strip()]

                address = (
                    row_data[headers["address"] - 1]
                    if "address" in headers and len(row_data) >= headers["address"]
                    else ""
                )
                addresses = [str(address).strip()] if address else []

                country = (
                    row_data[headers["country"] - 1]
                    if "country" in headers and len(row_data) >= headers["country"]
                    else ""
                )
                country = self._normalize_country_code(str(country or ""))

                license_req = (
                    row_data[headers["license_req"] - 1]
                    if "license_req" in headers and len(row_data) >= headers["license_req"]
                    else ""
                )
                license_policy = (
                    row_data[headers["license_policy"] - 1]
                    if "license_policy" in headers and len(row_data) >= headers["license_policy"]
                    else ""
                )
                fr_citation = (
                    row_data[headers["fr_citation"] - 1]
                    if "fr_citation" in headers and len(row_data) >= headers["fr_citation"]
                    else ""
                )

                effective_date = None
                if "effective_date" in headers and len(row_data) >= headers["effective_date"]:
                    date_val = row_data[headers["effective_date"] - 1]
                    if isinstance(date_val, datetime):
                        effective_date = date_val.date()
                    elif isinstance(date_val, date):
                        effective_date = date_val
                    elif isinstance(date_val, str):
                        effective_date = self._parse_date(date_val)

                entry = EntityListEntry(
                    id=f"EL-{row_idx:05d}",
                    name=str(name).strip(),
                    aliases=aliases,
                    addresses=addresses,
                    country=country,
                    license_requirement=str(license_req or ""),
                    license_policy=str(license_policy or ""),
                    federal_register_citation=str(fr_citation or ""),
                    effective_date=effective_date,
                )
                entries.append(entry)

            except Exception as e:
                logger.warning(f"Error parsing Entity List row {row_idx}: {e}")
                continue

        return entries

    def _normalize_country_code(self, country: str) -> str:
        """Normalize country name to ISO 3166-1 alpha-2 code."""
        country = country.strip().upper()

        # Common mappings
        mappings = {
            "CHINA": "CN",
            "PEOPLE'S REPUBLIC OF CHINA": "CN",
            "PRC": "CN",
            "RUSSIA": "RU",
            "RUSSIAN FEDERATION": "RU",
            "IRAN": "IR",
            "ISLAMIC REPUBLIC OF IRAN": "IR",
            "NORTH KOREA": "KP",
            "DPRK": "KP",
            "DEMOCRATIC PEOPLE'S REPUBLIC OF KOREA": "KP",
            "SYRIA": "SY",
            "SYRIAN ARAB REPUBLIC": "SY",
            "CUBA": "CU",
            "VENEZUELA": "VE",
            "BELARUS": "BY",
            "PAKISTAN": "PK",
            "UNITED ARAB EMIRATES": "AE",
            "UAE": "AE",
            "HONG KONG": "HK",
            "TAIWAN": "TW",
        }

        # If already a 2-letter code, return as-is
        if len(country) == 2:
            return country

        return mappings.get(country, country[:2] if len(country) >= 2 else "XX")

    # =========================================================================
    # Ingest All Sources
    # =========================================================================

    async def ingest_all(self, force_download: bool = False) -> dict[str, Any]:
        """
        Ingest all available sanctions lists.

        Args:
            force_download: If True, re-download all files.

        Returns:
            Dictionary with combined ingestion statistics.
        """
        results = {}

        # Clear existing data before full ingestion
        logger.info("Clearing existing sanctions data...")
        self.db.clear_all()

        # Ingest OFAC SDN
        logger.info("Ingesting OFAC SDN List...")
        results["ofac_sdn"] = await self.ingest_ofac_sdn(force_download)

        # Ingest BIS Denied Persons
        logger.info("Ingesting BIS Denied Persons List...")
        results["bis_denied_persons"] = await self.ingest_bis_denied_persons(force_download)

        # Note: Entity List requires manual download
        logger.info(
            "Note: BIS Entity List requires manual download from "
            "https://www.bis.doc.gov/index.php/policy-guidance/lists-of-parties-of-concern/entity-list"
        )

        # Summary
        total_added = sum(r.get("entries_added", 0) for r in results.values())
        total_errors = sum(len(r.get("errors", [])) for r in results.values())

        results["summary"] = {
            "total_entries_added": total_added,
            "total_errors": total_errors,
            "sources_processed": len(results) - 1,  # Exclude summary
        }

        return results
